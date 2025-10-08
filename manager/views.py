from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from django.db.models import Count, Sum, Q, Avg
from django.utils import timezone
from django.core.paginator import Paginator
from datetime import timedelta, date
from decimal import Decimal
from django.http import HttpResponse
from urllib.parse import urlencode
import uuid

from accounts.models import (
    User,
    Client,
    Domain,
    SubscriptionPlan,
    Subscription,
    Invoice,
    Payment,
    Announcement,
    SupportTicket,
    PlatformAnalytics,
)


def is_superadmin(user):
    """Check if user is superadmin"""
    return user.is_authenticated and user.role == User.Roles.SUPERADMIN


@login_required
@user_passes_test(is_superadmin)
def dashboard(request):
    """Main dashboard with key metrics"""
    today = timezone.now().date()

    # Basic counts
    total_tenants = Client.objects.count()
    active_tenants = Client.objects.filter(status=Client.Status.ACTIVE).count()
    trial_tenants = Client.objects.filter(status=Client.Status.TRIAL).count()
    suspended_tenants = Client.objects.filter(status=Client.Status.SUSPENDED).count()

    total_users = User.objects.count()
    total_plans = SubscriptionPlan.objects.filter(is_active=True).count()

    # Revenue metrics
    total_revenue = Payment.objects.filter(status=Payment.Status.COMPLETED).aggregate(
        total=Sum("amount")
    )["total"] or Decimal("0.00")

    monthly_revenue = Payment.objects.filter(
        status=Payment.Status.COMPLETED, created_at__gte=today.replace(day=1)
    ).aggregate(total=Sum("amount"))["total"] or Decimal("0.00")

    # Recent signups (last 30 days)
    recent_signups = Client.objects.filter(
        created_at__gte=timezone.now() - timedelta(days=30)
    ).count()

    # Recent tenants
    recent_tenants = Client.objects.order_by("-created_at")[:5]

    # Open support tickets
    open_tickets = SupportTicket.objects.filter(
        status__in=[SupportTicket.Status.OPEN, SupportTicket.Status.IN_PROGRESS]
    ).count()

    # Expiring subscriptions (next 7 days)
    expiring_soon = Subscription.objects.filter(
        expires_at__lte=today + timedelta(days=7),
        expires_at__gte=today,
        status=Subscription.Status.ACTIVE,
    ).count()

    # Monthly growth data (last 6 months)
    months_data = []
    for i in range(5, -1, -1):
        month_date = today - timedelta(days=30 * i)
        month_start = month_date.replace(day=1)
        if i > 0:
            next_month = (month_date + timedelta(days=32)).replace(day=1)
        else:
            next_month = today + timedelta(days=1)

        month_tenants = Client.objects.filter(
            created_at__gte=month_start, created_at__lt=next_month
        ).count()

        month_revenue = Payment.objects.filter(
            status=Payment.Status.COMPLETED,
            created_at__gte=month_start,
            created_at__lt=next_month,
        ).aggregate(total=Sum("amount"))["total"] or Decimal("0.00")

        months_data.append(
            {
                "month": month_start.strftime("%b %Y"),
                "tenants": month_tenants,
                "revenue": float(month_revenue),
            }
        )

    context = {
        "total_tenants": total_tenants,
        "active_tenants": active_tenants,
        "trial_tenants": trial_tenants,
        "suspended_tenants": suspended_tenants,
        "total_users": total_users,
        "total_plans": total_plans,
        "total_revenue": total_revenue,
        "monthly_revenue": monthly_revenue,
        "recent_signups": recent_signups,
        "recent_tenants": recent_tenants,
        "open_tickets": open_tickets,
        "expiring_soon": expiring_soon,
        "months_data": months_data,
        "now": timezone.now(),
    }

    return render(request, "manager/dashboard/dashboard.html", context)


@login_required
@user_passes_test(is_superadmin)
def tenant_list(request):
    """List all tenants with filtering"""
    tenants = Client.objects.prefetch_related("domains").all().order_by("-created_at")

    # Filtering
    status_filter = request.GET.get("status", "")
    search_query = request.GET.get("search", "")

    if status_filter:
        tenants = tenants.filter(status=status_filter)

    if search_query:
        tenants = tenants.filter(
            Q(name__icontains=search_query)
            | Q(schema_name__icontains=search_query)
            | Q(email__icontains=search_query)
        )
    # Pagination
    paginator = Paginator(tenants, 20)
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    context = {
        "page_obj": page_obj,
        "status_filter": status_filter,
        "search_query": search_query,
        "status_choices": Client.Status.choices,
    }

    return render(request, "manager/tenants/tenant_list.html", context)


@login_required
@user_passes_test(is_superadmin)
def tenant_detail(request, pk):
    """View tenant details"""
    tenant = get_object_or_404(Client, pk=pk)

    # Get related data
    domains = tenant.domains.all()
    users = User.objects.filter(groups__name=f"tenant_{tenant.schema_name}")
    subscription = getattr(tenant, "subscription", None)
    recent_invoices = Invoice.objects.filter(subscription__tenant=tenant).order_by(
        "-created_at"
    )[:5]

    context = {
        "tenant": tenant,
        "domains": domains,
        "users": users,
        "subscription": subscription,
        "recent_invoices": recent_invoices,
    }

    return render(request, "manager/tenants/tenant_detail.html", context)


@login_required
@user_passes_test(is_superadmin)
def tenant_create(request):
    """Create new tenant"""
    if request.method == "POST":
        name = request.POST.get("name")
        schema_name = request.POST.get("schema_name")
        email = request.POST.get("email")
        phone = request.POST.get("phone")
        address = request.POST.get("address")
        domain_name = request.POST.get("domain")

        try:
            # Create tenant
            tenant = Client.objects.create(
                name=name,
                schema_name=schema_name,
                email=email,
                phone=phone,
                address=address,
            )

            # Create domain
            Domain.objects.create(domain=domain_name, tenant=tenant, is_primary=True)

            # Create tenant owner (admin user)
            from django.db import connection

            # Switch to tenant schema to create the user
            connection.set_tenant(tenant)

            # Create admin user for the tenant
            owner_username = f"{schema_name}_admin"
            owner_email = email if email else f"{owner_username}@example.com"

            # Generate a random password
            import secrets
            import string

            alphabet = string.ascii_letters + string.digits
            random_password = "".join(secrets.choice(alphabet) for i in range(12))

            # Create the owner user
            owner = User.objects.create_user(
                username=owner_username,
                email=owner_email,
                password=random_password,
                role=User.Roles.ADMIN,
                first_name="Admin",
                last_name=name,
            )

            # Switch back to public schema
            connection.set_schema_to_public()

            messages.success(
                request,
                f'Tenant "{name}" created successfully! Owner username: {owner_username}, Password: {random_password} (Save this password!)',
            )
            return redirect("manager:tenant_detail", pk=tenant.pk)
        except Exception as e:
            messages.error(request, f"Error creating tenant: {str(e)}")

    return render(request, "manager/tenants/tenant_form.html", {"action": "Create"})


@login_required
@user_passes_test(is_superadmin)
def tenant_edit(request, pk):
    """Edit tenant"""
    tenant = get_object_or_404(Client, pk=pk)

    if request.method == "POST":
        tenant.name = request.POST.get("name", tenant.name)
        tenant.email = request.POST.get("email", tenant.email)
        tenant.phone = request.POST.get("phone", tenant.phone)
        tenant.address = request.POST.get("address", tenant.address)
        tenant.max_users = int(request.POST.get("max_users", tenant.max_users))
        tenant.max_products = int(request.POST.get("max_products", tenant.max_products))
        tenant.max_warehouses = int(
            request.POST.get("max_warehouses", tenant.max_warehouses)
        )

        try:
            tenant.save()
            messages.success(request, "Tenant updated successfully!")
            return redirect("manager:tenant_detail", pk=tenant.pk)
        except Exception as e:
            messages.error(request, f"Error updating tenant: {str(e)}")

    context = {"tenant": tenant, "action": "Edit"}

    return render(request, "manager/tenants/tenant_form.html", context)


@login_required
@user_passes_test(is_superadmin)
def tenant_delete(request, pk):
    """Delete tenant"""
    tenant = get_object_or_404(Client, pk=pk)

    if request.method == "POST":
        tenant_name = tenant.name
        try:
            tenant.delete()
            messages.success(request, f'Tenant "{tenant_name}" deleted successfully!')
            return redirect("manager:tenant_list")
        except Exception as e:
            messages.error(request, f"Error deleting tenant: {str(e)}")
            return redirect("manager:tenant_detail", pk=pk)

    context = {"tenant": tenant}
    return render(request, "manager/tenants/tenant_confirm_delete.html", context)


@login_required
@user_passes_test(is_superadmin)
def tenant_suspend(request, pk):
    """Suspend tenant"""
    tenant = get_object_or_404(Client, pk=pk)
    tenant.status = Client.Status.SUSPENDED
    tenant.save()
    messages.warning(request, f'Tenant "{tenant.name}" has been suspended.')
    return redirect("manager:tenant_detail", pk=pk)


@login_required
@user_passes_test(is_superadmin)
def tenant_activate(request, pk):
    """Activate tenant"""
    tenant = get_object_or_404(Client, pk=pk)
    tenant.status = Client.Status.ACTIVE
    tenant.save()
    messages.success(request, f'Tenant "{tenant.name}" has been activated.')
    return redirect("manager:tenant_detail", pk=pk)


@login_required
@user_passes_test(is_superadmin)
def user_list(request):
    """List all users"""
    users = User.objects.all().order_by("-date_joined")

    # Filtering
    role_filter = request.GET.get("role", "")
    search_query = request.GET.get("search", "")

    if role_filter:
        users = users.filter(role=role_filter)

    if search_query:
        users = users.filter(
            Q(username__icontains=search_query)
            | Q(email__icontains=search_query)
            | Q(first_name__icontains=search_query)
            | Q(last_name__icontains=search_query)
        )

    # Pagination
    paginator = Paginator(users, 20)
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    context = {
        "page_obj": page_obj,
        "role_filter": role_filter,
        "search_query": search_query,
        "role_choices": User.Roles.choices,
    }

    return render(request, "manager/users/user_list.html", context)


@login_required
@user_passes_test(is_superadmin)
def user_detail(request, pk):
    """View user details"""
    user = get_object_or_404(User, pk=pk)

    context = {
        "user_obj": user,
    }

    return render(request, "manager/users/user_detail.html", context)


@login_required
@user_passes_test(is_superadmin)
def user_create(request):
    """Create new user"""
    if request.method == "POST":
        username = request.POST.get("username")
        email = request.POST.get("email")
        password = request.POST.get("password")
        first_name = request.POST.get("first_name")
        last_name = request.POST.get("last_name")
        role = request.POST.get("role")
        phone = request.POST.get("phone")

        try:
            user = User.objects.create_user(
                username=username,
                email=email,
                password=password,
                first_name=first_name,
                last_name=last_name,
                role=role,
                phone=phone,
            )
            messages.success(request, f'User "{username}" created successfully!')
            return redirect("manager:user_detail", pk=user.pk)
        except Exception as e:
            messages.error(request, f"Error creating user: {str(e)}")

    context = {
        "action": "Create",
        "role_choices": User.Roles.choices,
    }

    return render(request, "manager/users/user_form.html", context)


@login_required
@user_passes_test(is_superadmin)
def user_edit(request, pk):
    """Edit user"""
    user = get_object_or_404(User, pk=pk)

    if request.method == "POST":
        user.email = request.POST.get("email", user.email)
        user.first_name = request.POST.get("first_name", user.first_name)
        user.last_name = request.POST.get("last_name", user.last_name)
        user.role = request.POST.get("role", user.role)
        user.phone = request.POST.get("phone", user.phone)
        user.is_active = request.POST.get("is_active") == "on"

        password = request.POST.get("password")
        if password:
            user.set_password(password)

        try:
            user.save()
            messages.success(request, "User updated successfully!")
            return redirect("manager:user_detail", pk=user.pk)
        except Exception as e:
            messages.error(request, f"Error updating user: {str(e)}")

    context = {
        "user_obj": user,
        "action": "Edit",
        "role_choices": User.Roles.choices,
    }

    return render(request, "manager/users/user_form.html", context)


@login_required
@user_passes_test(is_superadmin)
def user_delete(request, pk):
    """Delete user"""
    user = get_object_or_404(User, pk=pk)

    if request.method == "POST":
        username = user.username
        try:
            user.delete()
            messages.success(request, f'User "{username}" deleted successfully!')
            return redirect("manager:user_list")
        except Exception as e:
            messages.error(request, f"Error deleting user: {str(e)}")
            return redirect("manager:user_detail", pk=pk)

    context = {"user_obj": user}
    return render(request, "manager/users/user_confirm_delete.html", context)


@login_required
@user_passes_test(is_superadmin)
def plan_list(request):
    """List subscription plans"""
    plans = SubscriptionPlan.objects.all().order_by("plan_type")

    context = {"plans": plans}
    return render(request, "manager/plans/plan_list.html", context)


@login_required
@user_passes_test(is_superadmin)
def plan_create(request):
    """Create subscription plan"""
    if request.method == "POST":
        try:
            plan = SubscriptionPlan.objects.create(
                name=request.POST.get("name"),
                plan_type=request.POST.get("plan_type"),
                description=request.POST.get("description"),
                price_monthly=Decimal(request.POST.get("price_monthly")),
                price_yearly=Decimal(request.POST.get("price_yearly")),
                max_users=int(request.POST.get("max_users")),
                max_products=int(request.POST.get("max_products")),
                max_warehouses=int(request.POST.get("max_warehouses")),
                max_branches=int(request.POST.get("max_branches")),
                has_advanced_reporting=request.POST.get("has_advanced_reporting")
                == "on",
                has_api_access=request.POST.get("has_api_access") == "on",
                has_multi_currency=request.POST.get("has_multi_currency") == "on",
                has_offline_support=request.POST.get("has_offline_support") == "on",
            )
            messages.success(request, f'Plan "{plan.name}" created successfully!')
            return redirect("manager:plan_list")
        except Exception as e:
            messages.error(request, f"Error creating plan: {str(e)}")

    context = {
        "action": "Create",
        "plan_type_choices": SubscriptionPlan.PlanType.choices,
    }
    return render(request, "manager/plans/plan_form.html", context)


@login_required
@user_passes_test(is_superadmin)
def plan_edit(request, pk):
    """Edit subscription plan"""
    plan = get_object_or_404(SubscriptionPlan, pk=pk)

    if request.method == "POST":
        try:
            plan.name = request.POST.get("name")
            plan.description = request.POST.get("description")
            plan.price_monthly = Decimal(request.POST.get("price_monthly"))
            plan.price_yearly = Decimal(request.POST.get("price_yearly"))
            plan.max_users = int(request.POST.get("max_users"))
            plan.max_products = int(request.POST.get("max_products"))
            plan.max_warehouses = int(request.POST.get("max_warehouses"))
            plan.max_branches = int(request.POST.get("max_branches"))
            plan.has_advanced_reporting = (
                request.POST.get("has_advanced_reporting") == "on"
            )
            plan.has_api_access = request.POST.get("has_api_access") == "on"
            plan.has_multi_currency = request.POST.get("has_multi_currency") == "on"
            plan.has_offline_support = request.POST.get("has_offline_support") == "on"
            plan.is_active = request.POST.get("is_active") == "on"
            plan.save()

            messages.success(request, f'Plan "{plan.name}" updated successfully!')
            return redirect("manager:plan_list")
        except Exception as e:
            messages.error(request, f"Error updating plan: {str(e)}")

    context = {
        "plan": plan,
        "action": "Edit",
        "plan_type_choices": SubscriptionPlan.PlanType.choices,
    }
    return render(request, "manager/plans/plan_form.html", context)


@login_required
@user_passes_test(is_superadmin)
def plan_delete(request, pk):
    """Delete subscription plan"""
    plan = get_object_or_404(SubscriptionPlan, pk=pk)

    if request.method == "POST":
        plan_name = plan.name
        try:
            plan.delete()
            messages.success(request, f'Plan "{plan_name}" deleted successfully!')
        except Exception as e:
            messages.error(request, f"Error deleting plan: {str(e)}")
        return redirect("manager:plan_list")

    context = {"plan": plan}
    return render(request, "manager/plans/plan_confirm_delete.html", context)


@login_required
@user_passes_test(is_superadmin)
def subscription_list(request):
    """List all subscriptions"""
    subscriptions = Subscription.objects.select_related("tenant", "plan").order_by(
        "-created_at"
    )

    # Filtering
    status_filter = request.GET.get("status", "")
    if status_filter:
        subscriptions = subscriptions.filter(status=status_filter)

    # Pagination
    paginator = Paginator(subscriptions, 20)
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    context = {
        "page_obj": page_obj,
        "status_filter": status_filter,
        "status_choices": Subscription.Status.choices,
    }
    return render(request, "manager/plans/subscription_list.html", context)


@login_required
@user_passes_test(is_superadmin)
def subscription_detail(request, pk):
    """View subscription details"""
    subscription = get_object_or_404(Subscription, pk=pk)
    invoices = subscription.invoices.order_by("-created_at")[:10]
    payments = subscription.payments.order_by("-created_at")[:10]

    context = {
        "subscription": subscription,
        "invoices": invoices,
        "payments": payments,
    }
    return render(request, "manager/plans/subscription_detail.html", context)


@login_required
@user_passes_test(is_superadmin)
def subscription_create(request):
    """Create new subscription for a tenant"""
    if request.method == "POST":
        tenant_id = request.POST.get("tenant")
        plan_id = request.POST.get("plan")
        billing_cycle = request.POST.get("billing_cycle")
        auto_renew = request.POST.get("auto_renew") == "on"

        try:
            tenant = get_object_or_404(Client, pk=tenant_id)
            plan = get_object_or_404(SubscriptionPlan, pk=plan_id)

            # Check if tenant already has a subscription
            if hasattr(tenant, "subscription"):
                messages.error(
                    request,
                    f'Tenant "{tenant.name}" already has an active subscription.',
                )
                return redirect("manager:subscription_create")

            # Calculate expiry date based on billing cycle
            if billing_cycle == Subscription.BillingCycle.MONTHLY:
                expires_at = timezone.now() + timedelta(days=30)
            else:  # yearly
                expires_at = timezone.now() + timedelta(days=365)

            # Create subscription
            subscription = Subscription.objects.create(
                tenant=tenant,
                plan=plan,
                billing_cycle=billing_cycle,
                status=Subscription.Status.ACTIVE,
                expires_at=expires_at,
                auto_renew=auto_renew,
            )

            # Update tenant limits based on plan
            tenant.max_users = plan.max_users
            tenant.max_products = plan.max_products
            tenant.max_warehouses = plan.max_warehouses
            tenant.status = Client.Status.ACTIVE
            tenant.paid_until = expires_at.date()
            tenant.save()

            messages.success(
                request,
                f'Subscription created for "{tenant.name}" with {plan.name} plan.',
            )
            return redirect("manager:subscription_detail", pk=subscription.pk)
        except Exception as e:
            messages.error(request, f"Error creating subscription: {str(e)}")

    # Get tenants without subscriptions
    tenants_with_subs = Subscription.objects.values_list("tenant_id", flat=True)
    available_tenants = Client.objects.exclude(id__in=tenants_with_subs)
    plans = SubscriptionPlan.objects.filter(is_active=True)

    # Check if tenant is pre-selected via query param
    preselected_tenant_id = request.GET.get("tenant")
    preselected_tenant = None
    if preselected_tenant_id:
        try:
            preselected_tenant = Client.objects.get(pk=preselected_tenant_id)
            if hasattr(preselected_tenant, "subscription"):
                messages.warning(
                    request,
                    f'Tenant "{preselected_tenant.name}" already has a subscription.',
                )
                preselected_tenant = None
        except Client.DoesNotExist:
            pass

    context = {
        "action": "Create",
        "tenants": available_tenants,
        "plans": plans,
        "billing_cycle_choices": Subscription.BillingCycle.choices,
        "preselected_tenant": preselected_tenant,
    }
    return render(request, "manager/plans/subscription_form.html", context)


@login_required
@user_passes_test(is_superadmin)
def subscription_edit(request, pk):
    """Edit existing subscription"""
    subscription = get_object_or_404(Subscription, pk=pk)

    if request.method == "POST":
        plan_id = request.POST.get("plan")
        billing_cycle = request.POST.get("billing_cycle")
        auto_renew = request.POST.get("auto_renew") == "on"
        status = request.POST.get("status")

        try:
            # Update plan if changed
            if plan_id and int(plan_id) != subscription.plan.id:
                new_plan = get_object_or_404(SubscriptionPlan, pk=plan_id)
                subscription.plan = new_plan

                # Update tenant limits
                subscription.tenant.max_users = new_plan.max_users
                subscription.tenant.max_products = new_plan.max_products
                subscription.tenant.max_warehouses = new_plan.max_warehouses

                # Extend expiry date when plan is changed
                if subscription.expires_at and subscription.expires_at < timezone.now():
                    # If expired, set new expiry from now
                    if subscription.billing_cycle == Subscription.BillingCycle.MONTHLY:
                        subscription.expires_at = timezone.now() + timedelta(days=30)
                    else:
                        subscription.expires_at = timezone.now() + timedelta(days=365)
                elif subscription.expires_at:
                    # If still active, extend from current expiry
                    if subscription.billing_cycle == Subscription.BillingCycle.MONTHLY:
                        subscription.expires_at = subscription.expires_at + timedelta(
                            days=30
                        )
                    else:
                        subscription.expires_at = subscription.expires_at + timedelta(
                            days=365
                        )
                else:
                    # No expiry set, set from now
                    if subscription.billing_cycle == Subscription.BillingCycle.MONTHLY:
                        subscription.expires_at = timezone.now() + timedelta(days=30)
                    else:
                        subscription.expires_at = timezone.now() + timedelta(days=365)

                # Update tenant paid_until
                subscription.tenant.paid_until = subscription.expires_at.date()
                subscription.tenant.save()

            subscription.billing_cycle = billing_cycle
            subscription.auto_renew = auto_renew
            subscription.status = status
            subscription.save()

            messages.success(request, "Subscription updated successfully!")
            return redirect("manager:subscription_detail", pk=subscription.pk)
        except Exception as e:
            messages.error(request, f"Error updating subscription: {str(e)}")

    plans = SubscriptionPlan.objects.filter(is_active=True)

    context = {
        "action": "Edit",
        "subscription": subscription,
        "plans": plans,
        "billing_cycle_choices": Subscription.BillingCycle.choices,
        "status_choices": Subscription.Status.choices,
    }
    return render(request, "manager/plans/subscription_form.html", context)


@login_required
@user_passes_test(is_superadmin)
def subscription_cancel(request, pk):
    """Cancel a subscription"""
    subscription = get_object_or_404(Subscription, pk=pk)

    if request.method == "POST":
        subscription.status = Subscription.Status.CANCELLED
        subscription.cancelled_at = timezone.now()
        subscription.auto_renew = False
        subscription.save()

        # Update tenant status
        subscription.tenant.status = Client.Status.SUSPENDED
        subscription.tenant.save()

        messages.warning(
            request,
            f'Subscription for "{subscription.tenant.name}" has been cancelled.',
        )
        return redirect("manager:subscription_detail", pk=pk)

    context = {"subscription": subscription}
    return render(request, "manager/plans/subscription_confirm_cancel.html", context)


@login_required
@user_passes_test(is_superadmin)
def subscription_renew(request, pk):
    """Renew a subscription"""
    subscription = get_object_or_404(Subscription, pk=pk)

    if request.method == "POST":
        # Calculate new expiry date
        if subscription.billing_cycle == Subscription.BillingCycle.MONTHLY:
            new_expires_at = timezone.now() + timedelta(days=30)
        else:
            new_expires_at = timezone.now() + timedelta(days=365)

        subscription.status = Subscription.Status.ACTIVE
        subscription.expires_at = new_expires_at
        subscription.cancelled_at = None
        subscription.save()

        # Update tenant
        subscription.tenant.status = Client.Status.ACTIVE
        subscription.tenant.paid_until = new_expires_at.date()
        subscription.tenant.save()

        messages.success(
            request,
            f'Subscription for "{subscription.tenant.name}" has been renewed until {new_expires_at.date()}.',
        )
        return redirect("manager:subscription_detail", pk=pk)

    context = {"subscription": subscription}
    return render(request, "manager/plans/subscription_confirm_renew.html", context)


@login_required
@user_passes_test(is_superadmin)
def subscription_change_plan(request, pk):
    """Change subscription plan"""
    subscription = get_object_or_404(Subscription, pk=pk)

    if request.method == "POST":
        plan_id = request.POST.get("plan")

        try:
            new_plan = get_object_or_404(SubscriptionPlan, pk=plan_id)

            if new_plan == subscription.plan:
                messages.info(request, "Tenant is already on this plan.")
                return redirect("manager:subscription_detail", pk=pk)

            # Update subscription
            old_plan = subscription.plan
            subscription.plan = new_plan
            subscription.save()

            # Update tenant limits
            subscription.tenant.max_users = new_plan.max_users
            subscription.tenant.max_products = new_plan.max_products
            subscription.tenant.max_warehouses = new_plan.max_warehouses
            subscription.tenant.save()

            messages.success(
                request,
                f'Plan changed from "{old_plan.name}" to "{new_plan.name}" for "{subscription.tenant.name}".',
            )
            return redirect("manager:subscription_detail", pk=pk)
        except Exception as e:
            messages.error(request, f"Error changing plan: {str(e)}")

    plans = SubscriptionPlan.objects.filter(is_active=True).exclude(
        id=subscription.plan.id
    )

    context = {"subscription": subscription, "plans": plans}
    return render(request, "manager/plans/subscription_change_plan.html", context)


@login_required
@user_passes_test(is_superadmin)
def invoice_list(request):
    """List all invoices"""
    invoices = Invoice.objects.select_related(
        "subscription__tenant", "subscription__plan"
    ).order_by("-created_at")

    # Filtering
    status_filter = request.GET.get("status", "")
    subscription_id = request.GET.get("subscription", "")

    if status_filter:
        invoices = invoices.filter(status=status_filter)
    if subscription_id:
        invoices = invoices.filter(subscription_id=subscription_id)

    # Pagination
    paginator = Paginator(invoices, 20)
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    context = {
        "page_obj": page_obj,
        "status_filter": status_filter,
        "subscription_id": subscription_id,
        "status_choices": Invoice.Status.choices,
    }
    return render(request, "manager/invoices/invoice_list.html", context)


@login_required
@user_passes_test(is_superadmin)
def invoice_detail(request, pk):
    """View invoice details"""
    invoice = get_object_or_404(Invoice, pk=pk)
    payments = invoice.payments.order_by("-created_at")

    context = {
        "invoice": invoice,
        "payments": payments,
    }
    return render(request, "manager/invoices/invoice_detail.html", context)


@login_required
@user_passes_test(is_superadmin)
def invoice_export_excel(request):
    """Export invoices to Excel (xlsx) applying current filters"""
    # Build base queryset with same filters as list
    invoices = Invoice.objects.select_related(
        "subscription__tenant", "subscription__plan"
    ).order_by("-created_at")

    status_filter = request.GET.get("status", "")
    subscription_id = request.GET.get("subscription", "")

    if status_filter:
        invoices = invoices.filter(status=status_filter)
    if subscription_id:
        invoices = invoices.filter(subscription_id=subscription_id)

    # Generate Excel
    try:
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
    except Exception as e:
        messages.error(
            request,
            "Excel export dependencies are missing (openpyxl). Please install requirements.",
        )
        # Fallback to list view
        return redirect("manager:invoice_list")

    wb = Workbook()
    ws = wb.active
    ws.title = "Invoices"

    headers = [
        "Invoice #",
        "Tenant",
        "Plan",
        "Amount",
        "Currency",
        "Status",
        "Period Start",
        "Period End",
        "Due Date",
        "Paid At",
        "Created At",
    ]
    ws.append(headers)

    for inv in invoices:
        ws.append(
            [
                inv.invoice_number,
                inv.subscription.tenant.name,
                inv.subscription.plan.name,
                float(inv.amount),
                inv.currency,
                inv.get_status_display(),
                (
                    inv.billing_period_start.isoformat()
                    if inv.billing_period_start
                    else ""
                ),
                inv.billing_period_end.isoformat() if inv.billing_period_end else "",
                inv.due_date.isoformat() if inv.due_date else "",
                inv.paid_at.strftime("%Y-%m-%d %H:%M") if inv.paid_at else "",
                inv.created_at.strftime("%Y-%m-%d %H:%M"),
            ]
        )

    # Auto width
    for i, _ in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(i)].width = 18

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    filename = "invoices_export.xlsx"
    response["Content-Disposition"] = f"attachment; filename={filename}"
    wb.save(response)
    return response


@login_required
@user_passes_test(is_superadmin)
def invoice_export_pdf(request):
    """Export invoices to a simple PDF table applying current filters"""
    invoices = Invoice.objects.select_related(
        "subscription__tenant", "subscription__plan"
    ).order_by("-created_at")

    status_filter = request.GET.get("status", "")
    subscription_id = request.GET.get("subscription", "")

    if status_filter:
        invoices = invoices.filter(status=status_filter)
    if subscription_id:
        invoices = invoices.filter(subscription_id=subscription_id)

    try:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
        from reportlab.lib.styles import getSampleStyleSheet
    except Exception:
        messages.error(
            request,
            "PDF export dependencies are missing (reportlab). Please install requirements.",
        )
        return redirect("manager:invoice_list")

    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = "attachment; filename=invoices_export.pdf"

    doc = SimpleDocTemplate(
        response,
        pagesize=landscape(A4),
        leftMargin=24,
        rightMargin=24,
        topMargin=24,
        bottomMargin=24,
    )
    elements = []
    styles = getSampleStyleSheet()
    title = Paragraph("Invoices Export", styles["Title"])
    elements.append(title)

    data = [
        [
            "Invoice #",
            "Tenant",
            "Plan",
            "Amount",
            "Status",
            "Period",
            "Due",
            "Paid At",
            "Created",
        ]
    ]

    for inv in invoices:
        period = (
            f"{inv.billing_period_start} → {inv.billing_period_end}"
            if inv.billing_period_start and inv.billing_period_end
            else "-"
        )
        data.append(
            [
                inv.invoice_number,
                inv.subscription.tenant.name,
                inv.subscription.plan.name,
                f"{inv.currency} {inv.amount}",
                inv.get_status_display(),
                period,
                inv.due_date.isoformat() if inv.due_date else "-",
                inv.paid_at.strftime("%Y-%m-%d %H:%M") if inv.paid_at else "-",
                inv.created_at.strftime("%Y-%m-%d %H:%M"),
            ]
        )

    table = Table(data, repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.black),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("ALIGN", (3, 1), (3, -1), "RIGHT"),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
                (
                    "ROWBACKGROUNDS",
                    (0, 1),
                    (-1, -1),
                    [colors.whitesmoke, colors.lightyellow],
                ),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ]
        )
    )

    elements.append(table)
    doc.build(elements)
    return response


@login_required
@user_passes_test(is_superadmin)
def invoice_download_pdf(request, pk):
    """Generate a PDF for a single invoice"""
    invoice = get_object_or_404(Invoice, pk=pk)

    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.units import mm
        from reportlab.pdfgen import canvas
        from reportlab.lib import colors
    except Exception:
        messages.error(request, "PDF generation dependency missing (reportlab).")
        return redirect("manager:invoice_detail", pk=pk)

    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = (
        f"attachment; filename=invoice_{invoice.invoice_number}.pdf"
    )

    c = canvas.Canvas(response, pagesize=A4)
    width, height = A4

    # Header
    c.setFont("Helvetica-Bold", 18)
    c.drawString(20 * mm, height - 25 * mm, "INVOICE")
    c.setFont("Helvetica", 10)
    c.drawString(20 * mm, height - 32 * mm, f"Invoice #: {invoice.invoice_number}")
    c.drawString(
        20 * mm, height - 38 * mm, f"Date: {invoice.created_at.strftime('%Y-%m-%d')}"
    )

    # Tenant info
    tenant = invoice.subscription.tenant
    c.setFont("Helvetica-Bold", 12)
    c.drawString(20 * mm, height - 50 * mm, "Bill To:")
    c.setFont("Helvetica", 10)
    c.drawString(20 * mm, height - 56 * mm, tenant.name)
    if tenant.email:
        c.drawString(20 * mm, height - 62 * mm, tenant.email)
    if tenant.address:
        c.drawString(20 * mm, height - 68 * mm, tenant.address)

    # Summary box
    c.setStrokeColor(colors.grey)
    c.rect(120 * mm, height - 70 * mm, 70 * mm, 30 * mm, stroke=1, fill=0)
    c.setFont("Helvetica", 10)
    c.drawString(125 * mm, height - 50 * mm, f"Status: {invoice.get_status_display()}")
    c.drawString(
        125 * mm, height - 56 * mm, f"Amount: {invoice.currency} {invoice.amount}"
    )
    c.drawString(125 * mm, height - 62 * mm, f"Due: {invoice.due_date.isoformat()}")

    # Period
    c.setFont("Helvetica-Bold", 12)
    c.drawString(20 * mm, height - 85 * mm, "Billing Period")
    c.setFont("Helvetica", 10)
    c.drawString(
        20 * mm,
        height - 92 * mm,
        f"{invoice.billing_period_start.isoformat()} to {invoice.billing_period_end.isoformat()}",
    )

    # Notes
    if invoice.notes:
        c.setFont("Helvetica-Bold", 12)
        c.drawString(20 * mm, height - 110 * mm, "Notes")
        c.setFont("Helvetica", 10)
        text_obj = c.beginText(20 * mm, height - 117 * mm)
        for line in invoice.notes.splitlines():
            text_obj.textLine(line)
        c.drawText(text_obj)

    c.showPage()
    c.save()
    return response


@login_required
@user_passes_test(is_superadmin)
def payment_list(request):
    """List all payments with comprehensive filtering"""
    payments = Payment.objects.select_related(
        "subscription__tenant", "subscription__plan", "invoice"
    ).order_by("-created_at")

    # Filtering
    status_filter = request.GET.get("status", "")
    provider_filter = request.GET.get("provider", "")
    subscription_id = request.GET.get("subscription", "")
    invoice_id = request.GET.get("invoice", "")
    tenant_search = request.GET.get("tenant", "")

    if status_filter:
        payments = payments.filter(status=status_filter)
    if provider_filter:
        payments = payments.filter(provider=provider_filter)
    if subscription_id:
        payments = payments.filter(subscription_id=subscription_id)
    if invoice_id:
        payments = payments.filter(invoice_id=invoice_id)
    if tenant_search:
        payments = payments.filter(subscription__tenant__name__icontains=tenant_search)

    # Pagination
    paginator = Paginator(payments, 20)
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    context = {
        "page_obj": page_obj,
        "status_filter": status_filter,
        "provider_filter": provider_filter,
        "subscription_id": subscription_id,
        "invoice_id": invoice_id,
        "tenant_search": tenant_search,
        "status_choices": Payment.Status.choices,
        "provider_choices": Payment.Provider.choices,
    }
    return render(request, "manager/payments/payment_list.html", context)


@login_required
@user_passes_test(is_superadmin)
def payment_detail(request, pk):
    """View payment details with full context"""
    payment = get_object_or_404(Payment, pk=pk)

    context = {
        "payment": payment,
    }
    return render(request, "manager/payments/payment_detail.html", context)


@login_required
@user_passes_test(is_superadmin)
def payment_create(request):
    """Create a manual payment for a subscription"""
    if request.method == "POST":
        subscription_id = request.POST.get("subscription")
        amount = request.POST.get("amount")
        provider = request.POST.get("provider")
        notes = request.POST.get("notes", "")

        try:
            subscription = get_object_or_404(Subscription, pk=subscription_id)

            # Generate transaction ID
            transaction_id = (
                f"MAN-{timezone.now():%Y%m%d}-{uuid.uuid4().hex[:8].upper()}"
            )

            # Create payment
            payment = Payment.objects.create(
                subscription=subscription,
                provider=provider,
                transaction_id=transaction_id,
                amount=Decimal(amount),
                currency="USD",
                status=Payment.Status.COMPLETED,
                processed_at=timezone.now(),
                error_message=notes,
            )

            # Extend subscription expiry date
            if subscription.expires_at and subscription.expires_at < timezone.now():
                # If expired, set new expiry from now
                if subscription.billing_cycle == Subscription.BillingCycle.MONTHLY:
                    subscription.expires_at = timezone.now() + timedelta(days=30)
                else:
                    subscription.expires_at = timezone.now() + timedelta(days=365)
            elif subscription.expires_at:
                # If still active, extend from current expiry
                if subscription.billing_cycle == Subscription.BillingCycle.MONTHLY:
                    subscription.expires_at = subscription.expires_at + timedelta(
                        days=30
                    )
                else:
                    subscription.expires_at = subscription.expires_at + timedelta(
                        days=365
                    )
            else:
                # No expiry set, set from now
                if subscription.billing_cycle == Subscription.BillingCycle.MONTHLY:
                    subscription.expires_at = timezone.now() + timedelta(days=30)
                else:
                    subscription.expires_at = timezone.now() + timedelta(days=365)

            subscription.status = Subscription.Status.ACTIVE
            subscription.save()

            # Update tenant status
            tenant = subscription.tenant
            tenant.status = Client.Status.ACTIVE
            tenant.paid_until = subscription.expires_at.date()
            tenant.save()

            messages.success(
                request,
                f"Payment of ${amount} created successfully for {subscription.tenant.name}. Subscription extended to {subscription.expires_at.date()}.",
            )
            return redirect("manager:payment_detail", pk=payment.pk)
        except Exception as e:
            messages.error(request, f"Error creating payment: {str(e)}")

    # Get all active subscriptions
    subscriptions = Subscription.objects.select_related("tenant", "plan").order_by(
        "-created_at"
    )

    # Check if subscription is pre-selected via query param
    preselected_subscription_id = request.GET.get("subscription")
    preselected_subscription = None
    if preselected_subscription_id:
        try:
            preselected_subscription = Subscription.objects.get(
                pk=preselected_subscription_id
            )
        except Subscription.DoesNotExist:
            pass

    context = {
        "action": "Create",
        "subscriptions": subscriptions,
        "provider_choices": Payment.Provider.choices,
        "preselected_subscription": preselected_subscription,
    }
    return render(request, "manager/payments/payment_form.html", context)


@login_required
@user_passes_test(is_superadmin)
def payment_export_excel(request):
    """Export payments to Excel applying current filters"""
    payments = Payment.objects.select_related(
        "subscription__tenant", "subscription__plan", "invoice"
    ).order_by("-created_at")

    # Apply same filters as list view
    status_filter = request.GET.get("status", "")
    provider_filter = request.GET.get("provider", "")
    subscription_id = request.GET.get("subscription", "")
    invoice_id = request.GET.get("invoice", "")
    tenant_search = request.GET.get("tenant", "")

    if status_filter:
        payments = payments.filter(status=status_filter)
    if provider_filter:
        payments = payments.filter(provider=provider_filter)
    if subscription_id:
        payments = payments.filter(subscription_id=subscription_id)
    if invoice_id:
        payments = payments.filter(invoice_id=invoice_id)
    if tenant_search:
        payments = payments.filter(subscription__tenant__name__icontains=tenant_search)

    try:
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
    except Exception:
        messages.error(request, "Excel export requires openpyxl library.")
        return redirect("manager:payment_list")

    wb = Workbook()
    ws = wb.active
    ws.title = "Payments"

    headers = [
        "Transaction ID",
        "Tenant",
        "Plan",
        "Invoice #",
        "Provider",
        "Amount",
        "Currency",
        "Status",
        "Processed At",
        "Created At",
    ]
    ws.append(headers)

    for payment in payments:
        ws.append(
            [
                payment.transaction_id,
                payment.subscription.tenant.name,
                payment.subscription.plan.name,
                payment.invoice.invoice_number if payment.invoice else "N/A",
                payment.get_provider_display(),
                float(payment.amount),
                payment.currency,
                payment.get_status_display(),
                (
                    payment.processed_at.strftime("%Y-%m-%d %H:%M")
                    if payment.processed_at
                    else ""
                ),
                payment.created_at.strftime("%Y-%m-%d %H:%M"),
            ]
        )

    # Auto width
    for i, _ in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(i)].width = 18

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = "attachment; filename=payments_export.xlsx"
    wb.save(response)
    return response


@login_required
@user_passes_test(is_superadmin)
def payment_export_pdf(request):
    """Export payments to PDF applying current filters"""
    payments = Payment.objects.select_related(
        "subscription__tenant", "subscription__plan", "invoice"
    ).order_by("-created_at")

    # Apply same filters
    status_filter = request.GET.get("status", "")
    provider_filter = request.GET.get("provider", "")
    subscription_id = request.GET.get("subscription", "")
    invoice_id = request.GET.get("invoice", "")
    tenant_search = request.GET.get("tenant", "")

    if status_filter:
        payments = payments.filter(status=status_filter)
    if provider_filter:
        payments = payments.filter(provider=provider_filter)
    if subscription_id:
        payments = payments.filter(subscription_id=subscription_id)
    if invoice_id:
        payments = payments.filter(invoice_id=invoice_id)
    if tenant_search:
        payments = payments.filter(subscription__tenant__name__icontains=tenant_search)

    try:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
        from reportlab.lib.styles import getSampleStyleSheet
    except Exception:
        messages.error(request, "PDF export requires reportlab library.")
        return redirect("manager:payment_list")

    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = "attachment; filename=payments_export.pdf"

    doc = SimpleDocTemplate(
        response,
        pagesize=landscape(A4),
        leftMargin=24,
        rightMargin=24,
        topMargin=24,
        bottomMargin=24,
    )
    elements = []
    styles = getSampleStyleSheet()
    title = Paragraph("Payments Export", styles["Title"])
    elements.append(title)

    data = [
        [
            "Transaction",
            "Tenant",
            "Provider",
            "Amount",
            "Status",
            "Invoice",
            "Processed",
            "Created",
        ]
    ]

    for payment in payments:
        data.append(
            [
                payment.transaction_id[:20],
                payment.subscription.tenant.name[:20],
                payment.get_provider_display(),
                f"{payment.currency} {payment.amount}",
                payment.get_status_display(),
                payment.invoice.invoice_number if payment.invoice else "-",
                (
                    payment.processed_at.strftime("%Y-%m-%d %H:%M")
                    if payment.processed_at
                    else "-"
                ),
                payment.created_at.strftime("%Y-%m-%d %H:%M"),
            ]
        )

    table = Table(data, repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.black),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("ALIGN", (3, 1), (3, -1), "RIGHT"),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
                (
                    "ROWBACKGROUNDS",
                    (0, 1),
                    (-1, -1),
                    [colors.whitesmoke, colors.lightyellow],
                ),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ]
        )
    )

    elements.append(table)
    doc.build(elements)
    return response


@login_required
@user_passes_test(is_superadmin)
def announcement_list(request):
    """List all announcements"""
    announcements = Announcement.objects.order_by("-created_at")

    context = {"announcements": announcements}
    return render(request, "manager/announcements/announcement_list.html", context)


@login_required
@user_passes_test(is_superadmin)
def announcement_create(request):
    """Create announcement"""
    if request.method == "POST":
        try:
            announcement = Announcement.objects.create(
                title=request.POST.get("title"),
                content=request.POST.get("content"),
                priority=request.POST.get("priority"),
                is_active=request.POST.get("is_active") == "on",
                created_by=request.user,
            )

            # Add target tenants if specified
            tenant_ids = request.POST.getlist("target_tenants")
            if tenant_ids:
                announcement.target_tenants.set(tenant_ids)

            messages.success(request, "Announcement created successfully!")
            return redirect("manager:announcement_list")
        except Exception as e:
            messages.error(request, f"Error creating announcement: {str(e)}")

    context = {
        "action": "Create",
        "priority_choices": Announcement.Priority.choices,
        "tenants": Client.objects.all(),
    }
    return render(request, "manager/announcements/announcement_form.html", context)


@login_required
@user_passes_test(is_superadmin)
def announcement_edit(request, pk):
    """Edit announcement"""
    announcement = get_object_or_404(Announcement, pk=pk)

    if request.method == "POST":
        try:
            announcement.title = request.POST.get("title")
            announcement.content = request.POST.get("content")
            announcement.priority = request.POST.get("priority")
            announcement.is_active = request.POST.get("is_active") == "on"
            announcement.save()

            # Update target tenants
            tenant_ids = request.POST.getlist("target_tenants")
            announcement.target_tenants.set(tenant_ids if tenant_ids else [])

            messages.success(request, "Announcement updated successfully!")
            return redirect("manager:announcement_list")
        except Exception as e:
            messages.error(request, f"Error updating announcement: {str(e)}")

    context = {
        "announcement": announcement,
        "action": "Edit",
        "priority_choices": Announcement.Priority.choices,
        "tenants": Client.objects.all(),
    }
    return render(request, "manager/announcements/announcement_form.html", context)


@login_required
@user_passes_test(is_superadmin)
def announcement_delete(request, pk):
    """Delete announcement"""
    announcement = get_object_or_404(Announcement, pk=pk)

    if request.method == "POST":
        announcement.delete()
        messages.success(request, "Announcement deleted successfully!")
        return redirect("manager:announcement_list")

    context = {"announcement": announcement}
    return render(
        request, "manager/announcements/announcement_confirm_delete.html", context
    )


@login_required
@user_passes_test(is_superadmin)
def ticket_list(request):
    """List all support tickets"""
    tickets = SupportTicket.objects.select_related(
        "tenant", "created_by", "assigned_to"
    ).order_by("-created_at")

    # Filtering
    status_filter = request.GET.get("status", "")
    priority_filter = request.GET.get("priority", "")

    if status_filter:
        tickets = tickets.filter(status=status_filter)
    if priority_filter:
        tickets = tickets.filter(priority=priority_filter)

    # Pagination
    paginator = Paginator(tickets, 20)
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    context = {
        "page_obj": page_obj,
        "status_filter": status_filter,
        "priority_filter": priority_filter,
        "status_choices": SupportTicket.Status.choices,
        "priority_choices": SupportTicket.Priority.choices,
    }
    return render(request, "manager/tickets/ticket_list.html", context)


@login_required
@user_passes_test(is_superadmin)
def ticket_detail(request, pk):
    """View ticket details"""
    ticket = get_object_or_404(SupportTicket, pk=pk)

    context = {"ticket": ticket}
    return render(request, "manager/tickets/ticket_detail.html", context)


@login_required
@user_passes_test(is_superadmin)
def ticket_update_status(request, pk):
    """Update ticket status"""
    ticket = get_object_or_404(SupportTicket, pk=pk)

    if request.method == "POST":
        new_status = request.POST.get("status")
        ticket.status = new_status

        if new_status == SupportTicket.Status.RESOLVED:
            ticket.resolved_at = timezone.now()
        elif new_status == SupportTicket.Status.CLOSED:
            ticket.closed_at = timezone.now()

        ticket.save()
        messages.success(request, "Ticket status updated successfully!")

    return redirect("manager:ticket_detail", pk=pk)


@login_required
@user_passes_test(is_superadmin)
def analytics(request):
    """Platform analytics"""
    today = timezone.now().date()

    # Get or create today's analytics
    analytics_data, created = PlatformAnalytics.objects.get_or_create(
        date=today,
        defaults={
            "total_tenants": Client.objects.count(),
            "active_tenants": Client.objects.filter(
                status=Client.Status.ACTIVE
            ).count(),
            "trial_tenants": Client.objects.filter(status=Client.Status.TRIAL).count(),
            "total_users": User.objects.count(),
            "active_users": User.objects.filter(is_active=True).count(),
            "total_revenue": Payment.objects.filter(
                status=Payment.Status.COMPLETED
            ).aggregate(total=Sum("amount"))["total"]
            or Decimal("0.00"),
            "new_signups": Client.objects.filter(created_at__date=today).count(),
        },
    )

    # Historical data (last 30 days)
    historical_data = PlatformAnalytics.objects.filter(
        date__gte=today - timedelta(days=30)
    ).order_by("date")

    # Revenue by plan
    revenue_by_plan = (
        Subscription.objects.values("plan__name")
        .annotate(
            total_revenue=Sum(
                "payments__amount", filter=Q(payments__status=Payment.Status.COMPLETED)
            )
        )
        .order_by("-total_revenue")
    )

    # Tenant distribution by status
    tenant_distribution = Client.objects.values("status").annotate(count=Count("id"))

    context = {
        "analytics_data": analytics_data,
        "historical_data": historical_data,
        "revenue_by_plan": revenue_by_plan,
        "tenant_distribution": tenant_distribution,
    }
    return render(request, "manager/analytics.html", context)


@login_required
@user_passes_test(is_superadmin)
def reports(request):
    """Generate various reports"""
    today = timezone.now().date()

    # Date range filters
    start_date = request.GET.get("start_date", today - timedelta(days=30))
    end_date = request.GET.get("end_date", today)

    if isinstance(start_date, str):
        start_date = date.fromisoformat(start_date)
    if isinstance(end_date, str):
        end_date = date.fromisoformat(end_date)

    # Revenue report
    revenue_report = Payment.objects.filter(
        status=Payment.Status.COMPLETED,
        created_at__date__gte=start_date,
        created_at__date__lte=end_date,
    ).aggregate(total=Sum("amount"), count=Count("id"), avg=Avg("amount"))

    # Tenant growth
    tenant_growth = Client.objects.filter(
        created_at__date__gte=start_date, created_at__date__lte=end_date
    ).count()

    # Subscription changes
    new_subscriptions = Subscription.objects.filter(
        started_at__date__gte=start_date, started_at__date__lte=end_date
    ).count()

    cancelled_subscriptions = Subscription.objects.filter(
        cancelled_at__date__gte=start_date, cancelled_at__date__lte=end_date
    ).count()

    context = {
        "start_date": start_date,
        "end_date": end_date,
        "revenue_report": revenue_report,
        "tenant_growth": tenant_growth,
        "new_subscriptions": new_subscriptions,
        "cancelled_subscriptions": cancelled_subscriptions,
    }
    return render(request, "manager/reports.html", context)


# =============================================================================
# Authentication Views
# =============================================================================


def login_view(request):
    """Login page for admin panel"""
    from django.contrib.auth import authenticate, login

    if request.user.is_authenticated:
        # Already logged in
        if request.user.role == User.Roles.SUPERADMIN:
            return redirect("manager:dashboard")
        else:
            messages.error(request, "You do not have permission to access this area.")
            from django.contrib.auth import logout

            logout(request)

    if request.method == "POST":
        username = request.POST.get("username")
        password = request.POST.get("password")

        user = authenticate(request, username=username, password=password)

        if user is not None:
            if user.role == User.Roles.SUPERADMIN:
                login(request, user)
                next_url = request.GET.get("next", "/")
                return redirect(next_url)
            else:
                messages.error(
                    request, "You do not have permission to access this area."
                )
        else:
            messages.error(request, "Invalid username or password.")

    return render(request, "manager/login.html")


def logout_view(request):
    """Logout view for admin panel"""
    from django.contrib.auth import logout

    logout(request)
    messages.success(request, "You have been logged out successfully.")
    return redirect("manager:login")
